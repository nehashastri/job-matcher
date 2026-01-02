"""
Email notification module for job scraper
Sends email alerts for accepted jobs via Gmail SMTP
"""

import logging
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Dict, cast

import openpyxl
from config.config import DATA_DIR, Config
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class EmailNotifier:
    """Handles sending email notifications for matched jobs"""

    def __init__(self, config: Config):
        """
        Initialize email notifier

        Args:
            config: Configuration object with SMTP settings
        """
        self.config = config
        self.smtp_server = config.smtp_server
        self.smtp_port = config.smtp_port
        self.smtp_username = config.smtp_username
        self.smtp_password = config.smtp_password
        self.email_from = config.email_from
        self.email_to = config.email_to
        self.smtp_use_ssl = getattr(config, "smtp_use_ssl", False)
        self.enabled = config.enable_email_notifications
        self.connections_excel_path = Path(
            getattr(
                config,
                "connections_excel_path",
                DATA_DIR / "linkedin_connections.xlsx",
            )
        )

    def send_job_notification(
        self, job_data: Dict[str, Any], connection_count: int = 0
    ) -> bool:
        """
        Send email notification for a matched job

        Args:
            job_data: Dictionary containing job details
            connection_count: Number of connection requests sent

        Returns:
            bool: True if email sent successfully, False otherwise
        """
        if not self.enabled:
            logger.info("Email notifications disabled, skipping")
            return False

        if not self._validate_config():
            logger.error("Email configuration invalid, cannot send notification")
            return False

        try:
            role_contacts = self._load_role_contacts(
                job_data.get("title", ""), job_data.get("company", "")
            )

            # Compose email
            subject = self._compose_subject(job_data)
            body = self._compose_body(job_data, connection_count, role_contacts)

            # Create message
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = self.email_from
            msg["To"] = self.email_to

            # Attach plain text and HTML versions
            msg.attach(MIMEText(body, "plain"))
            msg.attach(
                MIMEText(
                    self._compose_html_body(job_data, connection_count, role_contacts),
                    "html",
                )
            )

            # Send email
            self._send_email(msg)

            logger.info(
                f"✅ Email notification sent for job: {job_data.get('title', 'Unknown')}"
            )
            return True

        except smtplib.SMTPAuthenticationError as e:
            logger.error(f"❌ SMTP authentication failed: {e}")
            logger.error("Check SMTP_USERNAME and SMTP_PASSWORD in .env")
            return False

        except smtplib.SMTPException as e:
            logger.error(f"❌ SMTP error occurred: {e}")
            return False

        except Exception as e:
            logger.error(f"❌ Unexpected error sending email: {e}")
            return False

    def _validate_config(self) -> bool:
        """
        Validate email configuration

        Returns:
            bool: True if config is valid, False otherwise
        """
        if not self.smtp_server:
            logger.error("SMTP_SERVER not configured in .env")
            return False
        if not self.smtp_username:
            logger.error("SMTP_USERNAME not configured in .env")
            return False
        if not self.smtp_password:
            logger.error("SMTP_PASSWORD not configured in .env")
            return False
        if not self.email_from:
            logger.error("EMAIL_FROM not configured in .env")
            return False
        if not self.email_to:
            logger.error("EMAIL_TO not configured in .env")
            return False
        return True

    def _send_email(self, msg: MIMEMultipart) -> None:
        """
        Send email via SMTP with TLS

        Args:
            msg: Email message to send

        Raises:
            smtplib.SMTPException: If SMTP error occurs
        """
        try:
            # Prefer SSL when configured or using port 465; otherwise use STARTTLS
            if self.smtp_use_ssl or self.smtp_port == 465:
                server = smtplib.SMTP_SSL(self.smtp_server, self.smtp_port, timeout=30)
                server.ehlo()
            else:
                server = smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=30)
                server.ehlo()
                server.starttls()
                server.ehlo()

            server.login(self.smtp_username, self.smtp_password)
            server.send_message(msg)
            server.quit()

        except smtplib.SMTPServerDisconnected as e:
            logger.error(f"❌ SMTP server disconnected: {e}")
            raise
        except smtplib.SMTPConnectError as e:
            logger.error(
                f"❌ Failed to connect to SMTP server {self.smtp_server}:{self.smtp_port}: {e}"
            )
            raise
        except Exception as e:
            logger.error(f"❌ Error during SMTP operation: {e}")
            raise

    def _compose_subject(self, job_data: Dict[str, Any]) -> str:
        """
        Compose email subject line

        Args:
            job_data: Job details

        Returns:
            str: Email subject
        """
        title = job_data.get("title", "Unknown")
        company = job_data.get("company", "Unknown")
        return f"🎯 Job Match: {title} at {company}"

    def _compose_body(
        self, job_data: Dict[str, Any], connection_count: int, role_contacts=None
    ) -> str:
        """
        Compose plain text email body

        Args:
            job_data: Job details
            connection_count: Number of connection requests sent
            role_contacts: Dict of role-specific contacts (messages/connections)

        Returns:
            str: Email body
        """
        title = job_data.get("title", "Unknown")
        company = job_data.get("company", "Unknown")
        location = job_data.get("location", "Unknown")
        match_score = job_data.get("match_score", 0)
        job_url = job_data.get("job_url", "")
        applicants = job_data.get("applicants", "Unknown")
        posted_date = job_data.get("posted_date", "Unknown")
        seniority = job_data.get("seniority", "Unknown")
        remote = job_data.get("remote", "Unknown")

        role_contacts = role_contacts or job_data.get("role_contacts") or {}
        message_targets = role_contacts.get("messages", [])
        connected_targets = role_contacts.get("connections", [])

        body = f"""
New Job Match Found!

Job Title: {title}
Company: {company}
Location: {location}
Match Score: {match_score}/10
Seniority: {seniority}
Remote: {remote}
Applicants: {applicants}
Posted: {posted_date}

Connection Requests Sent: {connection_count}

{self._render_contacts_text_section("Send Messages to", message_targets)}
{self._render_contacts_text_section("Connection Request sent to", connected_targets)}

Job URL: {job_url}

---
This is an automated notification from the Job Scraper.
"""
        return body

    def _compose_html_body(
        self, job_data: Dict[str, Any], connection_count: int, role_contacts=None
    ) -> str:
        """
        Compose HTML email body

        Args:
            job_data: Job details
            connection_count: Number of connection requests sent
            role_contacts: Dict of role-specific contacts (messages/connections)

        Returns:
            str: HTML email body
        """
        title = job_data.get("title", "Unknown")
        company = job_data.get("company", "Unknown")
        location = job_data.get("location", "Unknown")
        match_score = job_data.get("match_score", 0)
        job_url = job_data.get("job_url", "")
        applicants = job_data.get("applicants", "Unknown")
        posted_date = job_data.get("posted_date", "Unknown")
        seniority = job_data.get("seniority", "Unknown")
        remote = job_data.get("remote", "Unknown")

        role_contacts = role_contacts or {}
        message_targets = role_contacts.get("messages", [])
        connected_targets = role_contacts.get("connections", [])

        # Color code match score
        score_color = "#28a745" if match_score >= 8 else "#ffc107"

        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
        }}
        .header {{
            background-color: #0077b5;
            color: white;
            padding: 20px;
            text-align: center;
            border-radius: 5px 5px 0 0;
        }}
        .content {{
            background-color: #f8f9fa;
            padding: 20px;
            border: 1px solid #ddd;
            border-radius: 0 0 5px 5px;
        }}
        .job-title {{
            font-size: 24px;
            font-weight: bold;
            margin-bottom: 10px;
            color: #0077b5;
        }}
        .company {{
            font-size: 18px;
            color: #555;
            margin-bottom: 20px;
        }}
        .detail-row {{
            margin: 10px 0;
            padding: 8px;
            background-color: white;
            border-left: 3px solid #0077b5;
        }}
        .label {{
            font-weight: bold;
            display: inline-block;
            width: 140px;
        }}
        .match-score {{
            font-size: 20px;
            font-weight: bold;
            color: {score_color};
        }}
        .button {{
            display: inline-block;
            padding: 12px 24px;
            margin: 20px 0;
            background-color: #0077b5;
            color: white;
            text-decoration: none;
            border-radius: 5px;
            font-weight: bold;
        }}
        .footer {{
            margin-top: 20px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            font-size: 12px;
            color: #666;
            text-align: center;
        }}
        .section-title {{
            font-size: 18px;
            font-weight: bold;
            margin-top: 20px;
            color: #0077b5;
        }}
        .people-list {{
            list-style: none;
            padding: 0;
            margin: 10px 0 0;
        }}
        .people-list li {{
            background-color: white;
            margin-bottom: 8px;
            padding: 10px;
            border: 1px solid #e9ecef;
            border-radius: 4px;
        }}
        .people-name {{
            font-weight: bold;
            color: #004182;
        }}
        .people-title {{
            display: block;
            color: #555;
            margin-top: 4px;
            font-size: 14px;
        }}
        .people-link {{
            display: inline-block;
            margin-top: 6px;
            color: #0077b5;
            text-decoration: none;
            font-size: 13px;
        }}
        .people-link:hover {{
            text-decoration: underline;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🎯 New Job Match Found!</h1>
    </div>
    <div class="content">
        <div class="job-title">{title}</div>
        <div class="company">{company}</div>

        <div class="detail-row">
            <span class="label">Location:</span>
            <span>{location}</span>
        </div>

        <div class="detail-row">
            <span class="label">Match Score:</span>
            <span class="match-score">{match_score}/10</span>
        </div>

        <div class="detail-row">
            <span class="label">Seniority:</span>
            <span>{seniority}</span>
        </div>

        <div class="detail-row">
            <span class="label">Remote:</span>
            <span>{remote}</span>
        </div>

        <div class="detail-row">
            <span class="label">Applicants:</span>
            <span>{applicants}</span>
        </div>

        <div class="detail-row">
            <span class="label">Posted:</span>
            <span>{posted_date}</span>
        </div>

        <div class="detail-row">
            <span class="label">Connections Sent:</span>
            <span>{connection_count}</span>
        </div>

        {self._render_contacts_html_section("Send Messages to", message_targets)}
        {self._render_contacts_html_section("Connection Request sent to", connected_targets)}

        <a href="{job_url}" class="button">View Job on LinkedIn</a>

        <div class="footer">
            This is an automated notification from the Job Scraper.
        </div>
    </div>
</body>
</html>
"""
        return html

    def _render_contacts_text_section(
        self, title: str, contacts: list[dict[str, str]]
    ) -> str:
        if not contacts:
            return ""

        lines = [f"{title}:"]
        for person in contacts:
            name = person.get("name") or "Unknown"
            title_text = person.get("title") or ""
            url = person.get("url") or ""
            detail = f"{name}" if not title_text else f"{name} — {title_text}"
            if url:
                detail = f"{detail} ({url})"
            lines.append(f"- {detail}")
        return "\n".join(lines)

    def _render_contacts_html_section(
        self, title: str, contacts: list[dict[str, str]]
    ) -> str:
        if not contacts:
            return ""

        items = []
        for person in contacts:
            name = person.get("name") or "Unknown"
            title_text = person.get("title") or ""
            url = person.get("url") or ""
            link = (
                f'<a href="{url}" class="people-link" target="_blank">Profile</a>'
                if url
                else ""
            )
            items.append(
                f'<li><span class="people-name">{name}</span>'
                f'<span class="people-title">{title_text}</span>{link}</li>'
            )

        return (
            f'<div class="section-title">{title}</div>'
            f'<ul class="people-list">{"".join(items)}</ul>'
        )

    def _load_role_contacts(
        self, role: str, company: str | None = None
    ) -> dict[str, list[dict[str, str]]]:
        role_clean = (role or "").strip().lower()
        company_clean = (company or "").strip().lower()
        contacts = {"messages": [], "connections": []}

        if not role_clean:
            return contacts

        try:
            if not self.connections_excel_path.exists():
                logger.debug(
                    "No linkedin_connections.xlsx found at %s; skipping contact enrichment",
                    self.connections_excel_path,
                )
                return contacts

            wb = openpyxl.load_workbook(self.connections_excel_path)
            active_sheet = wb.active
            if active_sheet is None:
                return contacts
            ws = cast(Worksheet, active_sheet)
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            header_index = {name: idx for idx, name in enumerate(headers)}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                row_role = self._cell_text(
                    row[header_index.get("Role Searched", -1)]
                ).lower()
                if row_role != role_clean:
                    continue

                row_company = self._cell_text(
                    row[header_index.get("Company", -1)]
                ).lower()
                if company_clean and row_company and row_company != company_clean:
                    continue

                person = {
                    "name": self._cell_text(row[header_index.get("Name", -1)]),
                    "title": self._cell_text(row[header_index.get("Title", -1)]),
                    "url": self._cell_text(row[header_index.get("LinkedIn URL", -1)]),
                }

                message_available = (
                    self._cell_text(
                        row[header_index.get("Message Available", -1)]
                    ).lower()
                    == "yes"
                )
                connected = (
                    self._cell_text(row[header_index.get("Connected", -1)]).lower()
                    == "yes"
                )

                if message_available:
                    contacts["messages"].append(person)
                if connected:
                    contacts["connections"].append(person)

            return contacts

        except Exception as exc:
            logger.debug(f"Could not load contacts for role '{role}': {exc}")
            return contacts

    @staticmethod
    def _cell_text(value: Any) -> str:
        try:
            if value is None:
                return ""
            return str(value).strip()
        except Exception:
            return ""

    def test_connection(self) -> bool:
        """
        Test SMTP connection

        Returns:
            bool: True if connection successful, False otherwise
        """
        if not self._validate_config():
            return False

        try:
            logger.info(
                f"Testing SMTP connection to {self.smtp_server}:{self.smtp_port}..."
            )

            server = smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(self.smtp_username, self.smtp_password)
            server.quit()

            logger.info("✅ SMTP connection test successful")
            return True

        except smtplib.SMTPAuthenticationError as e:
            logger.error(f"❌ SMTP authentication failed: {e}")
            return False
        except smtplib.SMTPException as e:
            logger.error(f"❌ SMTP error: {e}")
            return False
        except Exception as e:
            logger.error(f"❌ Connection test failed: {e}")
            return False

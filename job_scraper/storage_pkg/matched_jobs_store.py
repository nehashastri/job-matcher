"""Job and contact storage backed by Excel (.xlsx)."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from config.config import DATA_DIR
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

JOBS_HEADERS = [
    "ID",
    "Title",
    "Company",
    "Location",
    "Job URL",
    "Source",
    "Applicants",
    "Posted Date",
    "Scraped Date",
    "Match Score",
    "Viewed",
    "Saved",
    "Applied",
    "Emailed",
]

CONNECTION_HEADERS = [
    "Date",
    "Name",
    "Title",
    "LinkedIn URL",
    "Company",
    "Country",
    "Role Searched",
    "Role Match",
    "Message Available",
    "Connected",
    "Status",
]


class MatchedJobsStore:
    """Store and manage jobs/contacts in Excel format."""

    def __init__(self, data_dir: str | Path | None = None):
        # Default to repo-level data directory when not provided
        self.data_dir = Path(data_dir) if data_dir else DATA_DIR
        self.data_dir.mkdir(exist_ok=True)

        self.jobs_excel = self.data_dir / "jobs.xlsx"
        self.connections_excel = self.data_dir / "linkedin_connections.xlsx"

        self._init_files()

    # ---------- Initialization ----------
    def _init_files(self) -> None:
        """Ensure Excel files exist."""
        if not self.jobs_excel.exists():
            self._write_jobs_excel([])
        if not self.connections_excel.exists():
            self._write_connections_excel([])

    # ---------- Jobs ----------
    def add_job(self, job: dict[str, Any]) -> bool:
        """Add or update a job in Excel."""
        try:
            jobs = self.get_all_jobs()
            job_id = str(
                job.get("id")
                or job.get("ID")
                or hash(job.get("url") or job.get("Job URL"))
            )

            title = job.get("title") or job.get("Title", "")
            company = job.get("company") or job.get("Company", "")
            location = job.get("location") or job.get("Location", "")
            job_url = job.get("url") or job.get("Job URL", "")
            source = job.get("source") or job.get("Source", "")
            applicants = job.get("applicant_count") or job.get("Applicants", 0)
            posted_date = job.get("posted_date") or job.get("Posted Date", "")
            match_score = job.get("match_score") or job.get("Match Score", 0)

            # Update if present
            for existing in jobs:
                if existing.get("ID") == job_id:
                    existing.update(
                        {
                            "Title": title,
                            "Company": company,
                            "Location": location,
                            "Job URL": job_url,
                            "Source": source,
                            "Applicants": applicants,
                            "Posted Date": posted_date,
                            "Match Score": match_score,
                        }
                    )
                    self._write_jobs_excel(jobs)
                    return True

            jobs.append(
                {
                    "ID": job_id,
                    "Title": title,
                    "Company": company,
                    "Location": location,
                    "Job URL": job_url,
                    "Source": source,
                    "Applicants": applicants,
                    "Posted Date": posted_date,
                    "Scraped Date": datetime.now().isoformat(),
                    "Match Score": match_score,
                    "Viewed": "No",
                    "Saved": "No",
                    "Applied": "No",
                    "Emailed": "No",
                }
            )

            self._write_jobs_excel(jobs)
            logger.info(f"Added job: {job.get('title')} at {job.get('company')}")
            return True
        except Exception as exc:
            logger.error(f"Error adding job: {exc}")
            return False

    def get_all_jobs(self) -> list[dict[str, Any]]:
        """Read all jobs from Excel."""
        if not self.jobs_excel.exists():
            return []
        try:
            wb = openpyxl.load_workbook(self.jobs_excel)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rows.append(
                    {
                        header: (row[idx] if row[idx] is not None else "")
                        for idx, header in enumerate(JOBS_HEADERS)
                    }
                )
            return rows
        except Exception as exc:
            logger.error(f"Error reading jobs.xlsx: {exc}")
            return []

    def _write_jobs_excel(self, jobs: list[dict[str, Any]]) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Jobs"
            ws.append(JOBS_HEADERS)

            for job in jobs:
                ws.append([job.get(h, "") for h in JOBS_HEADERS])

            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            widths = {
                "A": 12,
                "B": 30,
                "C": 24,
                "D": 18,
                "E": 44,
                "F": 16,
                "G": 14,
                "H": 18,
                "I": 22,
                "J": 14,
                "K": 10,
                "L": 10,
                "M": 10,
                "N": 10,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            for row_idx in range(2, len(jobs) + 2):
                url_cell = ws[f"E{row_idx}"]
                if url_cell.value:
                    url_cell.hyperlink = url_cell.value
                    url_cell.font = Font(color="0563C1", underline="single")

            wb.save(self.jobs_excel)
            return True
        except Exception as exc:
            logger.error(f"Error writing jobs.xlsx: {exc}")
            return False

    def export_to_excel(self, jobs: list[dict[str, Any]] | None = None) -> bool:
        """Public export hook (writes with formatting)."""
        jobs = jobs if jobs is not None else self.get_all_jobs()
        return self._write_jobs_excel(jobs)

    # ---------- Connections ----------
    def add_linkedin_connection(self, connection: dict[str, Any]) -> bool:
        """Record a LinkedIn connection attempt to Excel."""
        try:
            connections = self.get_all_connections()
            connections.append(
                {
                    "Date": datetime.now().isoformat(),
                    "Name": connection.get("name", ""),
                    "Title": connection.get("title", ""),
                    "LinkedIn URL": connection.get("url", ""),
                    "Company": connection.get("company", ""),
                    "Country": connection.get("country", ""),
                    "Role Searched": connection.get("role", ""),
                    "Role Match": "Yes"
                    if str(connection.get("role_match", "")).lower()
                    in {"true", "1", "yes"}
                    else "No",
                    "Message Available": "Yes"
                    if connection.get("message_available", False)
                    else "No",
                    "Connected": "Yes" if connection.get("connected", False) else "No",
                    "Status": connection.get("status", "Recorded"),
                }
            )

            self._write_connections_excel(connections)
            logger.info(f"Recorded connection to {connection.get('name')}")
            return True
        except Exception as exc:
            logger.error(f"Error adding connection: {exc}")
            return False

    def get_all_connections(self) -> list[dict[str, Any]]:
        """Read connections from Excel."""
        if not self.connections_excel.exists():
            return []
        try:
            wb = openpyxl.load_workbook(self.connections_excel)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rows.append(
                    {
                        header: (row[idx] if row[idx] is not None else "")
                        for idx, header in enumerate(CONNECTION_HEADERS)
                    }
                )
            return rows
        except Exception as exc:
            logger.error(f"Error reading linkedin_connections.xlsx: {exc}")
            return []

    def _write_connections_excel(self, connections: list[dict[str, Any]]) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "LinkedIn Connections"
            ws.append(CONNECTION_HEADERS)

            for conn in connections:
                ws.append([conn.get(h, "") for h in CONNECTION_HEADERS])

            header_fill = PatternFill(
                start_color="70AD47", end_color="70AD47", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            widths = {
                "A": 24,
                "B": 22,
                "C": 30,
                "D": 42,
                "E": 24,
                "F": 16,
                "G": 24,
                "H": 14,
                "I": 18,
                "J": 14,
                "K": 18,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            for row_idx in range(2, len(connections) + 2):
                url_cell = ws[f"D{row_idx}"]
                if url_cell.value:
                    url_cell.hyperlink = url_cell.value
                    url_cell.font = Font(color="0563C1", underline="single")

            wb.save(self.connections_excel)
            return True
        except Exception as exc:
            logger.error(f"Error writing linkedin_connections.xlsx: {exc}")
            return False

    def export_connections_to_excel(
        self, connections: list[dict[str, Any]] | None = None
    ) -> bool:
        connections = (
            connections if connections is not None else self.get_all_connections()
        )
        return self._write_connections_excel(connections)

    # ---------- Misc ----------
    def mark_job_status(self, job_id: str, status: str, value: bool = True):
        """Update job status (viewed, saved, applied, emailed)."""
        try:
            jobs = self.get_all_jobs()
            for job in jobs:
                if job.get("ID") == str(job_id):
                    job[status] = "Yes" if value else "No"
                    break
            self._write_jobs_excel(jobs)
            return True
        except Exception as exc:
            logger.error(f"Error updating job status: {exc}")
            return False

    def get_stats(self) -> dict[str, Any]:
        jobs = self.get_all_jobs()
        return {"total_jobs": len(jobs)}

    def _upgrade_connections_file(self):
        """Upgrade legacy linkedin_connections.csv to include role/company columns.

        Legacy files sometimes lacked Company/Role Match/Message/Connected columns.
        This method normalizes headers so downstream email linking can filter by role and company.
        """
        try:
            with open(self.connections_file, encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader) if reader else []

            expected_fields = {
                "Date": "",
                "Name": "",
                "Title": "",
                "LinkedIn URL": "",
                "Company": "",
                "Country": "",
                "Role Searched": "",
                "Role Match": "No",
                "Message Available": "No",
                "Connected": "No",
                "Status": "Recorded",
            }

            normalized: list[dict[str, Any]] = []
            for row in rows:
                normalized_row = expected_fields.copy()
                for key in normalized_row:
                    normalized_row[key] = row.get(key, normalized_row[key])

                # Legacy files used "Message Sent" to indicate outreach; map to Connected when yes.
                message_sent = (row.get("Message Sent") or "").strip().lower()
                if message_sent == "yes":
                    normalized_row["Connected"] = "Yes"

                # Legacy files may have company missing; keep blank when not present.
                normalized.append(normalized_row)

            if rows:
                self._write_connections_csv(normalized)

        except Exception as exc:
            logger.debug(f"Could not upgrade connections file: {exc}")
